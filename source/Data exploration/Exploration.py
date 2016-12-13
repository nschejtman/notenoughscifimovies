import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import itertools as itls
from openpyxl import load_workbook
from sklearn.tree import DecisionTreeClassifier
import sys
sys.path.append('./../')
import utils.utils as ut


def read_items():
    items_df = pd.read_csv('../../inputs/item_profile.csv', sep='\t')
    return items_df


def read_interactions():
    return pd.read_csv('../../inputs/interactions_idx.csv', sep='\t')


def write_description(data, ex_wrtr):
    print "Printing some statistics of the data\n"
    #print data.describe(include='all')
    try:
        data.describe(exclude=['object']).to_excel(ex_wrtr, sheet_name='Numerical_Description', na_rep='NA')
    except:
        pass
    try:
        data.describe(include=['object']).to_excel(ex_wrtr, sheet_name='Categorical_Description', na_rep='NA')
    except:
        pass
    ex_wrtr.save()


def plot_dist_num(X, name, note=None):
    ax = sns.distplot(X, kde = False, hist_kws={'alpha': 0.9})
    if note != None:
        plt.figtext(0, 0, note)
    ax.get_figure().savefig('./'+name+'.png', bbox_inches='tight')
    plt.close()




def feature_scorer(df, URM, criterion, filename):
    URM = ut.check_matrix(URM, format='csc')
    reps = np.diff(URM.indptr)
    X = np.repeat(df_copy.values, reps, axis=0)
    Y = URM.data
    clf = DecisionTreeClassifier(criterion=criterion)
    clf.fit(X, Y)
    pd.DataFrame(clf.feature_importances_, index=df_copy.columns, columns=['score']).sort_values('score', ascending=False).to_csv(filename + '.csv', sep='\t', header=False)



# Items dataframe description
# items_df = read_items()
# out_name = 'items.xlsx'
# ex_wrtr = pd.ExcelWriter(out_name, engine='openpyxl')
# try:
#     ex_wrtr.book = load_workbook(out_name)
#     ex_wrtr.sheets = dict((ws.title, ws) for ws in ex_wrtr.book.worksheets)
# except IOError:
#     pass
#
# write_description(items_df, ex_wrtr)

# Sum of ratings by item_id
# inters_df = pd.read_csv('../../inputs/interactions.csv', sep='\t')
# grp_itus = inters_df.groupby(['item_id', 'user_id'], as_index=False)['interaction_type'].agg(['max'])
# idxs = grp_itus.index.values
# aux = pd.DataFrame([[idxs[i][0], idxs[i][1], grp_itus.loc[idxs[i][0], idxs[i][1]]['max']] for i in range(len(idxs))],
#                       columns=['item_id', 'user_id', 'interaction_type'])
# grp_it = aux.groupby(['item_id'])['interaction_type'].agg(['sum'])
# grp_it.to_csv('../../inputs/interactions_sum.csv', sep='\t')

# Distribution of number of rated items
# inters_df = read_interactions()
# grp = inters_df.groupby(['user_idx'])['item_idx'].agg(['count']) #TODO: correct, count each item once
# plot_dist_num(grp['count'], "InteractionsPerUser_1", "All users")
# print "Users with more than 50 jobs: ", grp[grp['count'] > 50].shape[0]
# plot_dist_num(grp['count'][grp['count'] <= 50], "InteractionsPerUser_2",
#               "Users with no more than 50 jobs (" +str(grp[grp['count'] > 50].shape[0]) + " users excluded)")
# out_name = 'interactions.xlsx'
# ex_wrtr = pd.ExcelWriter(out_name, engine='openpyxl')
# try:
#     ex_wrtr.book = load_workbook(out_name)
#     ex_wrtr.sheets = dict((ws.title, ws) for ws in ex_wrtr.book.worksheets)
# except IOError:
#     pass
# write_description(grp, ex_wrtr)
# print grp[grp['count'] <= 2].shape[0]

urm = ut.read_interactions()
items_df = ut.read_items()
df_copy = pd.DataFrame(items_df, copy=True).drop(['id', 'latitude', 'longitude', 'created_at', 'active_during_test', 'title', 'tags'],axis=1)
country_dict = {c:i for i,c in enumerate(df_copy['country'].unique())}
df_copy['country'] = df_copy['country'].apply(lambda x: country_dict[x])
df_copy['country_reg'] = df_copy['country'].apply(lambda x: x*100) + df_copy['region']
df_copy = df_copy.drop(['country', 'region'], axis=1)
feature_scorer(items_df, urm, criterion='entropy', filename='Entropy feature importance 3')
feature_scorer(items_df, urm, criterion='gini', filename='Gini feature importance 3')